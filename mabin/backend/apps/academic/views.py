from rest_framework import viewsets, status, serializers, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.utils import timezone
from django.db.models import Q, Avg, Count, Sum, F
from django.db import transaction
from django_filters.rest_framework import DjangoFilterBackend
from django.contrib.auth import get_user_model
from openpyxl import load_workbook
from apps.approval.models import LeaveApplication
from apps.rbac.models import Department
from .models import Course, CourseSelection, Grade, StudentStatus, AttendanceSession, Attendance, Assignment, AssignmentSubmission
from .serializers import (
    CourseSerializer, CourseSelectionSerializer, GradeSerializer,
    StudentStatusSerializer, AttendanceSessionSerializer, AttendanceSerializer,
    AssignmentSerializer, AssignmentSubmissionSerializer
)

User = get_user_model()


class CourseViewSet(viewsets.ModelViewSet):
    """课程视图集"""
    queryset = Course.objects.filter(is_active=True)
    serializer_class = CourseSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['course_type', 'semester', 'department', 'teacher']
    search_fields = ['course_code', 'name', 'description']
    ordering_fields = ['created_at', 'course_code', 'current_students']

    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': '未上传文件'}, status=status.HTTP_400_BAD_REQUEST)

        filename = file_obj.name.lower()
        if not (filename.endswith('.xls') or filename.endswith('.xlsx')):
            return Response({'error': '仅支持 .xls 或 .xlsx 文件'}, status=status.HTTP_400_BAD_REQUEST)

        if not (request.user.is_staff or request.user.is_superuser or getattr(request.user, 'is_secretary', False)):
            return Response({'error': '无权导入课程数据'}, status=status.HTTP_403_FORBIDDEN)

        try:
            workbook = load_workbook(file_obj, data_only=True)
        except Exception:
            return Response({'error': 'Excel 文件解析失败'}, status=status.HTTP_400_BAD_REQUEST)

        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            return Response({'error': 'Excel 内容为空'}, status=status.HTTP_400_BAD_REQUEST)

        headers = []
        for cell in rows[0]:
            headers.append(str(cell).strip() if cell is not None else '')

        def get_index(aliases):
            for alias in aliases:
                if alias in headers:
                    return headers.index(alias)
            return None

        idx_code = get_index(['课程代码', 'course_code', '代码'])
        idx_name = get_index(['课程名称', 'name', '课程名'])
        idx_credit = get_index(['学分', 'credit'])
        idx_hours = get_index(['学时', 'hours'])
        idx_type = get_index(['课程类型', '类型', 'course_type'])
        idx_teacher = get_index(['授课教师', '教师', 'teacher'])
        idx_semester = get_index(['学期', 'semester'])
        idx_max_students = get_index(['最大选课人数', '容量', 'max_students'])
        idx_assessment = get_index(['考核方式', 'assessment_method'])
        idx_department = get_index(['开课学院', '学院', 'department'])

        if idx_code is None or idx_name is None:
            return Response({'error': '表头中必须包含“课程代码”和“课程名称”列'}, status=status.HTTP_400_BAD_REQUEST)

        type_map = {
            '必修': 'required',
            '必修课': 'required',
            'required': 'required',
            '选修': 'elective',
            '选修课': 'elective',
            'elective': 'elective',
            '公选': 'public',
            '公选课': 'public',
            '公共选修': 'public',
            'public': 'public',
        }

        created_count = 0
        updated_count = 0
        errors = []

        with transaction.atomic():
            for index, row in enumerate(rows[1:], start=2):
                if not row:
                    continue

                raw_code = row[idx_code] if idx_code is not None and idx_code < len(row) else None
                raw_name = row[idx_name] if idx_name is not None and idx_name < len(row) else None

                course_code = str(raw_code).strip() if raw_code is not None else ''
                name = str(raw_name).strip() if raw_name is not None else ''

                if not course_code or not name:
                    errors.append(f'第 {index} 行课程代码或课程名称为空，已跳过')
                    continue

                credit = None
                hours = None
                course_type = None
                teacher = None
                semester = ''
                max_students = None
                assessment_method = ''
                department = None

                if idx_credit is not None and idx_credit < len(row):
                    value = row[idx_credit]
                    if value is not None:
                        try:
                            credit = float(value)
                        except (TypeError, ValueError):
                            errors.append(f'第 {index} 行学分格式不正确，已使用默认值')

                if idx_hours is not None and idx_hours < len(row):
                    value = row[idx_hours]
                    if value is not None:
                        try:
                            hours = int(value)
                        except (TypeError, ValueError):
                            errors.append(f'第 {index} 行学时格式不正确，已使用默认值')

                if idx_type is not None and idx_type < len(row):
                    value = row[idx_type]
                    if value is not None:
                        type_raw = str(value).strip()
                        course_type = type_map.get(type_raw, None)
                        if course_type is None and type_raw:
                            if type_raw in ['required', 'elective', 'public']:
                                course_type = type_raw
                            else:
                                errors.append(f'第 {index} 行课程类型“{type_raw}”无法识别，已使用默认值')

                if idx_semester is not None and idx_semester < len(row):
                    value = row[idx_semester]
                    if value is not None:
                        semester = str(value).strip()

                if idx_max_students is not None and idx_max_students < len(row):
                    value = row[idx_max_students]
                    if value is not None:
                        try:
                            max_students = int(value)
                        except (TypeError, ValueError):
                            errors.append(f'第 {index} 行最大选课人数格式不正确，已使用默认值')

                if idx_assessment is not None and idx_assessment < len(row):
                    value = row[idx_assessment]
                    if value is not None:
                        assessment_method = str(value).strip()

                if idx_department is not None and idx_department < len(row):
                    value = row[idx_department]
                    if value is not None:
                        dept_name = str(value).strip()
                        if dept_name:
                            dept = Department.objects.filter(name=dept_name).first()
                            if not dept:
                                errors.append(f'第 {index} 行开课学院“{dept_name}”不存在，已忽略')
                            else:
                                department = dept

                if idx_teacher is not None and idx_teacher < len(row):
                    value = row[idx_teacher]
                    if value is not None:
                        teacher_key = str(value).strip()
                        if teacher_key:
                            teacher = User.objects.filter(username=teacher_key).first()
                            if not teacher:
                                teacher = User.objects.filter(real_name=teacher_key).first()
                            if not teacher:
                                errors.append(f'第 {index} 行授课教师“{teacher_key}”未找到，已忽略')

                course_defaults = {
                    'name': name,
                }
                if credit is not None:
                    course_defaults['credit'] = credit
                if hours is not None:
                    course_defaults['hours'] = hours
                if course_type:
                    course_defaults['course_type'] = course_type
                if semester:
                    course_defaults['semester'] = semester
                if max_students is not None:
                    course_defaults['max_students'] = max_students
                if department:
                    course_defaults['department'] = department
                if teacher:
                    course_defaults['teacher'] = teacher
                if assessment_method:
                    course_defaults['assessment_method'] = assessment_method

                course, created = Course.objects.get_or_create(
                    course_code=course_code,
                    defaults=course_defaults,
                )

                if created:
                    created_count += 1
                else:
                    updated = False
                    for field, value in course_defaults.items():
                        if value not in [None, '']:
                            setattr(course, field, value)
                            updated = True
                    if updated:
                        course.save()
                        updated_count += 1

        return Response(
            {
                'created': created_count,
                'updated': updated_count,
                'errors': errors,
            }
        )

    def get_queryset(self):
        """根据用户角色过滤"""
        user = self.request.user
        queryset = super().get_queryset().select_related('teacher', 'department')
        
        if user.is_superuser:
            return queryset
            
        # 教师查看自己授课的课程
        if user.is_teacher:
            return queryset.filter(teacher=user)
            
        # 学生查看自己已选的课程
        if getattr(user, 'is_student_user', False) and hasattr(user, 'student_profile'):
            return queryset.filter(
                selections__student=user.student_profile,
                selections__status='selected'
            ).distinct()
            
        # 学院秘书/领导查看本学院的课程
        if (user.is_secretary or user.has_role('leader')) and user.department:
            return queryset.filter(department=user.department)
            
        return queryset.none()
    
    @action(detail=False, methods=['get'])
    def available(self, request):
        """获取可选课程列表（供学生选课使用）"""
        # 选课列表不应受 get_queryset 的角色限制
        queryset = Course.objects.filter(is_active=True).select_related('teacher', 'department')
        semester = request.query_params.get('semester')
        
        # 过滤未满员的课程
        queryset = queryset.filter(
            current_students__lt=F('max_students')
        )
        
        if semester:
            queryset = queryset.filter(semester=semester)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def students(self, request, pk=None):
        """获取选课学生列表"""
        course = self.get_object()
        selections = CourseSelection.objects.filter(
            course=course,
            status='selected'
        )
        serializer = CourseSelectionSerializer(selections, many=True)
        return Response(serializer.data)


class CourseSelectionViewSet(viewsets.ModelViewSet):
    """选课视图集"""
    queryset = CourseSelection.objects.all()
    serializer_class = CourseSelectionSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """根据用户角色过滤"""
        user = self.request.user
        queryset = super().get_queryset()
        
        if user.is_superuser:
            return queryset
        
        # 学生只能看自己的选课
        if hasattr(user, 'student_profile'):
            return queryset.filter(student=user.student_profile)
        
        # 教师可以看自己课程的选课
        return queryset.filter(course__teacher=user)
    
    @transaction.atomic
    def perform_create(self, serializer):
        """选课"""
        course = serializer.validated_data['course']
        student = serializer.validated_data['student']
        
        # 检查是否已选
        if CourseSelection.objects.filter(student=student, course=course, status='selected').exists():
            raise serializers.ValidationError('该课程已选')
        
        # 检查是否已满
        if course.current_students >= course.max_students:
            raise serializers.ValidationError('该课程已满')
        
        # 创建选课记录
        selection = serializer.save(status='selected')
        
        # 更新课程选课人数
        course.current_students += 1
        course.save()
    
    @action(detail=True, methods=['post'])
    def drop(self, request, pk=None):
        """退课"""
        selection = self.get_object()
        
        if selection.status != 'selected':
            return Response({'error': '该课程未在选课状态'}, status=status.HTTP_400_BAD_REQUEST)
        
        selection.status = 'dropped'
        selection.dropped_at = timezone.now()
        selection.save()
        
        # 更新课程选课人数
        selection.course.current_students -= 1
        selection.course.save()
        
        serializer = self.get_serializer(selection)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def my_courses(self, request):
        """获取我的选课"""
        if not hasattr(request.user, 'student_profile'):
            return Response({'error': '非学生用户'}, status=status.HTTP_400_BAD_REQUEST)
        
        selections = CourseSelection.objects.filter(
            student=request.user.student_profile,
            status='selected'
        )
        serializer = self.get_serializer(selections, many=True)
        return Response(serializer.data)


class GradeViewSet(viewsets.ModelViewSet):
    """成绩视图集"""
    queryset = Grade.objects.all()
    serializer_class = GradeSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['semester', 'course', 'student']
    search_fields = ['student__user__real_name', 'student__student_id', 'course__name', 'course__course_code']
    ordering_fields = ['score', 'grade_point', 'recorded_at']
    
    def get_queryset(self):
        """根据用户角色过滤"""
        user = self.request.user
        queryset = super().get_queryset()
        
        if user.is_superuser:
            return queryset
        
        # 学生只能看自己的成绩
        if user.is_student_user and hasattr(user, 'student_profile'):
            return queryset.filter(student=user.student_profile)
        
        # 教师可以录入和查看自己课程的成绩
        if user.is_teacher:
            return queryset.filter(course__teacher=user)
        
        # 学院秘书可以查看本学院学生的成绩
        if user.is_secretary and user.department:
            return queryset.filter(student__user__department=user.department)
        
        # 院领导可以查看本学院所有成绩
        if user.has_role('leader') and user.department:
            return queryset.filter(student__user__department=user.department)
            
        return queryset.none()
    
    def perform_create(self, serializer):
        """录入成绩时自动计算绩点和等级"""
        grade = serializer.save(teacher=self.request.user)
        
        if grade.score:
            grade.grade_point = grade.calculate_grade_point()
            
            # 计算等级
            score = float(grade.score)
            if score >= 95:
                grade.grade_level = 'A+'
            elif score >= 90:
                grade.grade_level = 'A'
            elif score >= 85:
                grade.grade_level = 'A-'
            elif score >= 80:
                grade.grade_level = 'B+'
            elif score >= 75:
                grade.grade_level = 'B'
            elif score >= 70:
                grade.grade_level = 'B-'
            elif score >= 65:
                grade.grade_level = 'C+'
            elif score >= 60:
                grade.grade_level = 'C'
            else:
                grade.grade_level = 'F'
            
            grade.save()
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """成绩统计"""
        semester = request.query_params.get('semester')
        
        queryset = self.get_queryset()
        if semester:
            queryset = queryset.filter(semester=semester)
            
        if not queryset.exists():
            return Response({
                'average_score': 0,
                'average_gpa': 0,
                'total_credits': 0,
                'total_courses': 0,
                'pass_rate': 0
            })

        # 计算平均分
        avg_score = queryset.aggregate(Avg('score'))['score__avg'] or 0
        
        # 计算平均绩点 (GPA)
        avg_gpa = queryset.aggregate(Avg('grade_point'))['grade_point__avg'] or 0
        
        # 计算已获学分 (及格的课程)
        total_credits = queryset.filter(score__gte=60).aggregate(Sum('course__credit'))['course__credit__sum'] or 0
        
        # 总课程数
        total_courses = queryset.count()
        
        # 及格率
        pass_count = queryset.filter(score__gte=60).count()
        pass_rate = (pass_count / total_courses * 100) if total_courses > 0 else 0
        
        return Response({
            'average_score': round(avg_score, 2),
            'average_gpa': round(avg_gpa, 2),
            'total_credits': total_credits,
            'total_courses': total_courses,
            'pass_rate': round(pass_rate, 2)
        })

    @action(detail=False, methods=['get'])
    def semesters(self, request):
        """获取所有有成绩记录的学期列表"""
        semesters = Grade.objects.values_list('semester', flat=True).distinct().order_by('-semester')
        return Response(list(semesters))


class StudentStatusViewSet(viewsets.ModelViewSet):
    """学籍状态视图集"""
    queryset = StudentStatus.objects.all()
    serializer_class = StudentStatusSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """根据用户角色过滤"""
        user = self.request.user
        queryset = super().get_queryset()
        
        if user.is_superuser:
            return queryset
        
        # 学生只能看自己的学籍
        if hasattr(user, 'student_profile'):
            return queryset.filter(student=user.student_profile)
        
        # 学院秘书可以查看本学院学生的学籍
        if user.department:
            return queryset.filter(student__user__department=user.department)
        
        return queryset.none()
    
    @action(detail=True, methods=['post'])
    def change_status(self, request, pk=None):
        """变更学籍状态"""
        status_obj = self.get_object()
        new_status = request.data.get('status')
        reason = request.data.get('reason', '')
        
        if not new_status:
            return Response({'error': '必须指定新状态'}, status=status.HTTP_400_BAD_REQUEST)
        
        # 记录状态变更
        status_obj.status_changes.append({
            'from_status': status_obj.status,
            'to_status': new_status,
            'reason': reason,
            'operator': request.user.real_name,
            'changed_at': timezone.now().isoformat()
        })
        
        status_obj.status = new_status
        status_obj.save()
        
        serializer = self.get_serializer(status_obj)
        return Response(serializer.data)


class AttendanceSessionViewSet(viewsets.ModelViewSet):
    """签到会话视图集"""
    queryset = AttendanceSession.objects.all()
    serializer_class = AttendanceSessionSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['course', 'is_active']
    ordering_fields = ['created_at']

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        if user.is_superuser:
            return queryset
        # 教师看自己创建的
        if hasattr(user, 'taught_courses'):
            return queryset.filter(teacher=user)
        # 学生看自己选课的
        if hasattr(user, 'student_profile'):
            return queryset.filter(
                course__selections__student=user.student_profile,
                course__selections__status='selected'
            )
        return queryset.none()

    def perform_create(self, serializer):
        # 只有授课教师可以创建签到
        course = serializer.validated_data['course']
        if not self.request.user.is_superuser and course.teacher != self.request.user:
            raise serializers.ValidationError("只有授课教师可以发起签到")
        
        # 自动生成签到码 (4位数字)
        import random
        check_in_code = str(random.randint(1000, 9999))
        serializer.save(teacher=self.request.user, check_in_code=check_in_code)

    @action(detail=True, methods=['post'])
    def close(self, request, pk=None):
        """关闭签到会话"""
        session = self.get_object()
        session.is_active = False
        session.save()
        return Response({'message': '签到已关闭'})


class AttendanceViewSet(viewsets.ModelViewSet):
    """课堂签到视图集"""
    queryset = Attendance.objects.all()
    serializer_class = AttendanceSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['course', 'student', 'date', 'status', 'session']
    search_fields = ['student__user__real_name', 'course__name']
    ordering_fields = ['date', 'created_at']
    
    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        
        if user.is_superuser:
            return queryset
            
        # 学生只能看自己的签到
        if hasattr(user, 'student_profile'):
            return queryset.filter(student=user.student_profile)
            
        # 教师可以看自己课程的签到
        if hasattr(user, 'taught_courses'):
             return queryset.filter(course__teacher=user)
             
        return queryset.none()
    
    @action(detail=False, methods=['post'])
    def check_in(self, request):
        """学生签到"""
        session_id = request.data.get('session_id')
        code = request.data.get('code')
        
        if not session_id or not code:
            return Response({'error': '缺少参数'}, status=400)
            
        try:
            session = AttendanceSession.objects.get(id=session_id)
            if not session.is_active:
                return Response({'error': '签到已关闭'}, status=400)
            
            if session.expire_at and session.expire_at < timezone.now():
                session.is_active = False
                session.save()
                return Response({'error': '签到已过期'}, status=400)
                
            if session.check_in_code != code:
                return Response({'error': '签到码错误'}, status=400)
            
            student = request.user.student_profile
            # 检查是否已选该课
            if not CourseSelection.objects.filter(student=student, course=session.course, status='selected').exists():
                return Response({'error': '您未选修该课程'}, status=400)
                
            # 记录签到
            attendance, created = Attendance.objects.update_or_create(
                student=student,
                course=session.course,
                date=session.date,
                time_slot=session.time_slot,
                defaults={
                    'session': session,
                    'status': 'present'
                }
            )
            
            return Response({
                'message': '签到成功',
                'data': AttendanceSerializer(attendance).data
            })
        except AttendanceSession.DoesNotExist:
            return Response({'error': '签到会话不存在'}, status=404)
        except Exception as e:
            return Response({'error': str(e)}, status=400)
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """签到统计"""
        student_id = request.query_params.get('student_id')
        course_id = request.query_params.get('course_id')
        
        queryset = self.get_queryset()
        if student_id:
            queryset = queryset.filter(student_id=student_id)
        if course_id:
            queryset = queryset.filter(course_id=course_id)
            
        total = queryset.count()
        if total == 0:
            return Response({
                'total': 0,
                'present_rate': 0,
                'status_counts': {}
            })
            
        status_counts = queryset.values('status').annotate(count=Count('id'))
        present_count = queryset.filter(status='present').count()
        
        return Response({
            'total': total,
            'present_count': present_count,
            'present_rate': round(present_count / total * 100, 2),
            'status_counts': {item['status']: item['count'] for item in status_counts}
        })

    @action(detail=False, methods=['get'])
    def check_leaves(self, request):
        """批量检查学生在指定日期是否有批准的请假"""
        student_ids = request.query_params.getlist('student_ids[]')
        date_str = request.query_params.get('date')
        
        if not student_ids or not date_str:
            return Response({'error': '缺少参数'}, status=400)
            
        try:
            date = timezone.datetime.strptime(date_str, '%Y-%m-%d').date()
            # 查找在指定日期有批准请假的学生
            # 请假开始时间 <= 该日 23:59:59 且 请假结束时间 >= 该日 00:00:00
            start_of_day = timezone.make_aware(timezone.datetime.combine(date, timezone.datetime.min.time()))
            end_of_day = timezone.make_aware(timezone.datetime.combine(date, timezone.datetime.max.time()))
            
            leaves = LeaveApplication.objects.filter(
                applicant__student_profile__id__in=student_ids,
                status='approved',
                start_time__lte=end_of_day,
                end_time__gte=start_of_day
            ).values_list('applicant__student_profile__id', flat=True)
            
            return Response({'leave_student_ids': list(leaves)})
        except Exception as e:
            return Response({'error': str(e)}, status=400)

    @action(detail=False, methods=['post'])
    def batch_record(self, request):
        """批量录入签到"""
        course_id = request.data.get('course_id')
        date = request.data.get('date', timezone.now().date())
        time_slot = request.data.get('time_slot', '')
        records = request.data.get('records', [])  # [{"student_id": 1, "status": "present"}]
        
        if not course_id or not records:
            return Response({'error': '参数不全'}, status=status.HTTP_400_BAD_REQUEST)
            
        try:
            with transaction.atomic():
                for item in records:
                    Attendance.objects.update_or_create(
                        student_id=item['student_id'],
                        course_id=course_id,
                        date=date,
                        time_slot=time_slot,
                        defaults={'status': item['status']}
                    )
            return Response({'message': '录入成功'})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class AssignmentViewSet(viewsets.ModelViewSet):
    """作业视图集"""
    queryset = Assignment.objects.all()
    serializer_class = AssignmentSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['course']
    search_fields = ['title', 'description']
    ordering_fields = ['deadline', 'created_at']

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()

        if user.is_superuser:
            return queryset

        # 教师查看自己授课的作业
        if hasattr(user, 'taught_courses'):
            return queryset.filter(course__teacher=user)

        # 学生查看自己选课的作业
        if hasattr(user, 'student_profile'):
            return queryset.filter(
                course__selections__student=user.student_profile,
                course__selections__status='selected'
            )

        return queryset.none()


    def perform_create(self, serializer):
        user = self.request.user
        # 只有课程教师或管理员可以创建作业
        course = serializer.validated_data.get('course')
        if not user.is_superuser and course.teacher != user:
            raise serializers.ValidationError("您没有权限为该课程发布作业")
        serializer.save()

class AssignmentSubmissionViewSet(viewsets.ModelViewSet):
    """作业提交视图集"""
    queryset = AssignmentSubmission.objects.all()
    serializer_class = AssignmentSubmissionSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['assignment', 'student', 'status']
    search_fields = ['student__user__real_name', 'assignment__title']
    ordering_fields = ['submitted_at', 'score']

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()

        if user.is_superuser:
            return queryset

        # 教师查看自己发布的作业提交
        if hasattr(user, 'taught_courses'):
            return queryset.filter(assignment__course__teacher=user)

        # 学生只能看自己的提交
        if hasattr(user, 'student_profile'):
            return queryset.filter(student=user.student_profile)

        return queryset.none()

    def perform_create(self, serializer):
        user = self.request.user
        if not hasattr(user, 'student_profile'):
            raise serializers.ValidationError("只有学生可以提交作业")
        
        assignment = serializer.validated_data['assignment']
        now = timezone.now()
        is_late = now > assignment.deadline
        
        serializer.save(
            student=user.student_profile,
            is_late=is_late,
            submitted_at=now
        )

    @action(detail=True, methods=['post'])
    def grade(self, request, pk=None):
        """教师评分"""
        submission = self.get_object()
        
        # 权限检查：只有课程教师或管理员可以评分
        if not request.user.is_superuser and submission.assignment.course.teacher != request.user:
            return Response({'error': '无权评分'}, status=status.HTTP_403_FORBIDDEN)
            
        score = request.data.get('score')
        feedback = request.data.get('feedback', '')
        
        if score is None:
            return Response({'error': '分数不能为空'}, status=status.HTTP_400_BAD_REQUEST)
            
        try:
            score_val = float(score)
            if not (0 <= score_val <= 100):
                return Response({'error': '分数必须在0-100之间'}, status=status.HTTP_400_BAD_REQUEST)
        except ValueError:
            return Response({'error': '非法分数格式'}, status=status.HTTP_400_BAD_REQUEST)
            
        submission.score = score
        submission.feedback = feedback
        submission.status = 'graded'
        submission.save()
        
        # 自动同步到 Grade 表 (可选，根据业务逻辑决定是否实时同步)
        # 这里建议手动或在特定流程下同步，避免数据不一致
        
        serializer = self.get_serializer(submission)
        return Response(serializer.data)
