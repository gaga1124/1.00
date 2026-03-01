from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from django.contrib.auth import get_user_model
from django.db import transaction
from openpyxl import load_workbook
from .models import Student, PoliticalStatusRecord, StudentRecord, StudentStatusChange, ArchiveViewLog
from .serializers import (
    StudentSerializer, PoliticalStatusRecordSerializer, StudentRecordSerializer,
    StudentStatusChangeSerializer, ArchiveViewLogSerializer
)

User = get_user_model()


class StudentViewSet(viewsets.ModelViewSet):
    """学生视图集"""
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['major', 'class_name', 'grade', 'political_status', 'department', 'user__gender']
    search_fields = ['student_id', 'user__real_name', 'user__username', 'major', 'class_name']
    ordering_fields = ['created_at', 'student_id']
    ordering = ['-created_at']

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        # 记录查阅日志
        if request.user != instance.user:
            ArchiveViewLog.objects.create(
                student=instance,
                viewer=request.user,
                reason=request.query_params.get('view_reason', '常规查阅'),
                ip_address=request.META.get('REMOTE_ADDR')
            )
        
        serializer = self.get_serializer(instance, context={'include_details': True})
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def records(self, request, pk=None):
        """获取学生所有档案记录"""
        student = self.get_object()
        records = student.records.all()
        serializer = StudentRecordSerializer(records, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def political_history(self, request, pk=None):
        """获取学生政治面貌流转历史"""
        student = self.get_object()
        records = student.political_status_records.all()
        serializer = PoliticalStatusRecordSerializer(records, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def update_political_status(self, request, pk=None):
        """更新学生政治面貌（触发审批流程）"""
        student = self.get_object()
        to_status = request.data.get('to_status')
        
        if not to_status:
            return Response({'error': '必须指定目标状态'}, status=400)
        
        # 创建政治面貌流转记录
        record = PoliticalStatusRecord.objects.create(
            student=student,
            from_status=student.political_status,
            to_status=to_status,
            application_date=request.data.get('application_date'),
            application_file=request.data.get('application_file'),
            thought_report=request.data.get('thought_report'),
        )
        
        # TODO: 触发审批流程
        
        serializer = PoliticalStatusRecordSerializer(record)
        return Response(serializer.data, status=201)

    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': '未上传文件'}, status=status.HTTP_400_BAD_REQUEST)
        filename = file_obj.name.lower()
        if not (filename.endswith('.xls') or filename.endswith('.xlsx')):
            return Response({'error': '仅支持 .xls 或 .xlsx 文件'}, status=status.HTTP_400_BAD_REQUEST)
        if not (request.user.is_staff or request.user.is_superuser or getattr(request.user, 'is_secretary', False)):
            return Response({'error': '无权导入学生数据'}, status=status.HTTP_403_FORBIDDEN)
        try:
            workbook = load_workbook(file_obj, data_only=True)
        except Exception:
            return Response({'error': 'Excel 文件解析失败'}, status=status.HTTP_400_BAD_REQUEST)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            return Response({'error': 'Excel 内容为空'}, status=status.HTTP_400_BAD_REQUEST)
        headers = []
        for cell in rows[0]:
            headers.append(str(cell).strip() if cell is not None else '')

        def get_index(aliases):
            for alias in aliases:
                if alias in headers:
                    return headers.index(alias)
            return None

        idx_student_id = get_index(['学号', 'student_id'])
        idx_name = get_index(['姓名', 'name', 'real_name'])
        idx_major = get_index(['专业', 'major'])
        idx_class = get_index(['班级', 'class_name', '班级名称'])
        idx_grade = get_index(['年级', 'grade'])
        idx_political = get_index(['政治面貌', 'political_status'])
        idx_phone = get_index(['手机号', '电话', 'phone'])
        idx_email = get_index(['邮箱', 'email'])
        idx_id_card = get_index(['身份证号', '身份证', 'id_card', 'idcard'])

        if idx_student_id is None or idx_name is None:
            return Response({'error': '表头中必须包含“学号”和“姓名”列'}, status=status.HTTP_400_BAD_REQUEST)

        political_map = {
            '群众': 'masses',
            '团员': 'member',
            '入党积极分子': 'activist',
            '预备党员': 'probationary',
            '正式党员': 'party_member',
            'masses': 'masses',
            'member': 'member',
            'activist': 'activist',
            'probationary': 'probationary',
            'party_member': 'party_member',
        }

        created_count = 0
        updated_count = 0
        errors = []

        with transaction.atomic():
            for index, row in enumerate(rows[1:], start=2):
                if not row:
                    continue
                raw_student_id = row[idx_student_id] if idx_student_id is not None else None
                raw_name = row[idx_name] if idx_name is not None else None
                student_id = str(raw_student_id).strip() if raw_student_id is not None else ''
                real_name = str(raw_name).strip() if raw_name is not None else ''
                if not student_id or not real_name:
                    errors.append(f'第 {index} 行学号或姓名为空，已跳过')
                    continue
                major = ''
                class_name = ''
                grade = ''
                political_display = ''
                phone = ''
                email = ''
                id_card = ''
                if idx_major is not None and idx_major < len(row):
                    value = row[idx_major]
                    major = str(value).strip() if value is not None else ''
                if idx_class is not None and idx_class < len(row):
                    value = row[idx_class]
                    class_name = str(value).strip() if value is not None else ''
                if idx_grade is not None and idx_grade < len(row):
                    value = row[idx_grade]
                    grade = str(value).strip() if value is not None else ''
                if idx_political is not None and idx_political < len(row):
                    value = row[idx_political]
                    political_display = str(value).strip() if value is not None else ''
                if idx_phone is not None and idx_phone < len(row):
                    value = row[idx_phone]
                    phone = str(value).strip() if value is not None else ''
                if idx_email is not None and idx_email < len(row):
                    value = row[idx_email]
                    email = str(value).strip() if value is not None else ''
                if idx_id_card is not None and idx_id_card < len(row):
                    value = row[idx_id_card]
                    id_card = str(value).strip() if value is not None else ''

                political_status = political_map.get(political_display, 'masses')

                user, user_created = User.objects.get_or_create(
                    username=student_id,
                    defaults={
                        'real_name': real_name,
                        'phone': phone,
                        'id_card': id_card,
                        'email': email,
                    },
                )
                if user_created:
                    user.set_password(student_id)
                    user.save()
                else:
                    changed = False
                    if real_name and not user.real_name:
                        user.real_name = real_name
                        changed = True
                    if phone and not user.phone:
                        user.phone = phone
                        changed = True
                    if email and not user.email:
                        user.email = email
                        changed = True
                    if id_card and not getattr(user, 'id_card', None):
                        user.id_card = id_card
                        changed = True
                    if changed:
                        user.save()

                student, stu_created = Student.objects.get_or_create(
                    student_id=student_id,
                    defaults={
                        'user': user,
                        'major': major,
                        'class_name': class_name,
                        'grade': grade,
                        'political_status': political_status,
                    },
                )

                if stu_created:
                    created_count += 1
                else:
                    updated = False
                    if major:
                        student.major = major
                        updated = True
                    if class_name:
                        student.class_name = class_name
                        updated = True
                    if grade:
                        student.grade = grade
                        updated = True
                    if political_status:
                        student.political_status = political_status
                        updated = True
                    if student.user_id != user.id:
                        student.user = user
                        updated = True
                    if updated:
                        student.save()
                        updated_count += 1

        return Response(
            {
                'created': created_count,
                'updated': updated_count,
                'errors': errors,
            }
        )


class PoliticalStatusRecordViewSet(viewsets.ModelViewSet):
    """政治面貌流转记录视图集"""
    queryset = PoliticalStatusRecord.objects.all()
    serializer_class = PoliticalStatusRecordSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['student', 'status', 'from_status', 'to_status']
    search_fields = ['student__student_id', 'student__user__real_name']
    ordering_fields = ['created_at', 'application_date']
    ordering = ['-created_at']

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        if user.is_superuser:
            return queryset
        
        # 学生只能看到自己的记录
        if hasattr(user, 'student_profile'):
            return queryset.filter(student=user.student_profile)
        
        # 教师/辅导员可以看到本学院学生或所有（根据需求，这里暂定全看，后续可加学院过滤）
        return queryset

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """审批政治面貌流转"""
        record = self.get_object()
        action_type = request.data.get('action')  # 'approve' or 'reject'
        comment = request.data.get('comment', '')
        
        if action_type == 'approve':
            record.status = 'approved'
            record.approval_date = request.data.get('approval_date', timezone.now().date())
            # 更新学生政治面貌
            record.student.political_status = record.to_status
            record.student.save()
        elif action_type == 'reject':
            record.status = 'rejected'
        else:
            return Response({'error': '无效的操作类型'}, status=400)
        
        record.approver = request.user
        record.approval_comment = comment
        record.save()
        
        serializer = self.get_serializer(record)
        return Response(serializer.data)


class StudentStatusChangeViewSet(viewsets.ModelViewSet):
    """学籍变动记录视图集"""
    queryset = StudentStatusChange.objects.all()
    serializer_class = StudentStatusChangeSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['student', 'change_type']
    search_fields = ['student__student_id', 'student__user__real_name', 'document_no', 'reason']
    ordering_fields = ['change_date', 'created_at']
    ordering = ['-change_date']

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        if user.is_superuser:
            return queryset
        
        if hasattr(user, 'student_profile'):
            return queryset.filter(student=user.student_profile)
        
        return queryset
    
    def perform_create(self, serializer):
        serializer.save(operator=self.request.user)


class ArchiveViewLogViewSet(viewsets.ReadOnlyModelViewSet):
    """档案查阅日志视图集"""
    queryset = ArchiveViewLog.objects.all()
    serializer_class = ArchiveViewLogSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['student', 'viewer']
    search_fields = ['student__user__real_name', 'viewer__real_name']
    ordering_fields = ['viewed_at']
    ordering = ['-viewed_at']

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        if user.is_superuser:
            return queryset
        
        # 只有管理员或相关负责人可以查看所有日志
        # 学生只能看到谁查阅了自己的档案
        if hasattr(user, 'student_profile'):
            return queryset.filter(student=user.student_profile)
        
        return queryset


class StudentRecordViewSet(viewsets.ModelViewSet):
    """学生档案记录视图集"""
    queryset = StudentRecord.objects.all()
    serializer_class = StudentRecordSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['student', 'record_type']
    search_fields = ['title', 'description', 'student__student_id', 'student__user__real_name']
    ordering_fields = ['created_at']
    ordering = ['-created_at']

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        if user.is_superuser:
            return queryset
        
        if hasattr(user, 'student_profile'):
            return queryset.filter(student=user.student_profile)
        
        return queryset
    
    def perform_create(self, serializer):
        """创建记录时自动设置操作人"""
        serializer.save(operator=self.request.user)
